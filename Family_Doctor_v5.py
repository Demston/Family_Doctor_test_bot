""" Family_Doctor_test_bot.
Тестовый бот условной медицинской клиники. Позволяет вести запись клиентов через Телеграм, удобен и прост.
В папке с файлом бота лежат файлы *.xlsx, разрешение должно быть именно такое, в файле должен быть включен общий доступ.
Имя файла - профессия доктора, имя листа - дата (чч.мм.гггг), данные в таблице пополняются по мере записи клиентов.
Файлы необходимо проверять и корректировать вручную (добавлять/удалять листы с новыми/старыми датами).
Если в штате появился доктор с новой профессией - нужно просто создать новый файл, доктор отобразится в боте.
Имена всех инлайновых кнопок бота подтягиваются из файла."""

import openpyxl
import telebot
from telebot import types
import os

TOKEN = '6937469925:AAEnHeyx5EnupxzicTUl1EoJ4SGXb9PYZ5Y'
bot = telebot.TeleBot(TOKEN)

user_data = {'formal_dict': {}}
# словарь с параметрами конкретного клиента, ключ - id его чата


@bot.message_handler(commands=['start'])
def welcome(message):
    """Начало работы с ботом"""
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    button1 = types.KeyboardButton('Записаться к доктору 🗓️')
    button2 = types.KeyboardButton('О нас 🏥')
    markup.add(button1, button2)
    bot.send_message(message.chat.id, '{0.first_name}, клиника "Фэмили Доктор" рада приветствовать Вас !'
                     .format(message.from_user, bot.get_me()), parse_mode='html',
                     reply_markup=markup)


@bot.message_handler(content_types=['text'])
def dialog(message):
    """Работа с текстовыми сообщениями"""
    global user_data
    if message.chat.type == 'private':
        if message.text == 'О нас 🏥':
            bot.send_message(message.chat.id, 'ООО "Фэмили Доктор", ИНН 3567778888, '
                                              'Юр. адрес г. Москва, м. Пушкина, ул. Колотушкина, д. 228')
        elif message.text == 'Записаться к доктору 🗓️':
            markup = types.InlineKeyboardMarkup()
            # есть папка, в папке - файлы, имена файлов - доктора, они подтянутся в инлайновые кнопки
            files_names = [each.split('.xls')[0]
                           for each in os.listdir() if each.endswith('.xlsx')]
            # цикл для создания кнопок с выбором доктора
            for i in files_names:
                markup.add((types.InlineKeyboardButton(str(i), callback_data=str(i))))
            bot.send_message(message.chat.id, 'К какому доктору Вы хотели бы записаться?', reply_markup=markup)
        # эта часть срабатывает, когда клиент уже заполнил параметры: доктор, дата, время (функция callback_doctor)
        elif user_data[message.chat.id]['time'] is not None:
            if ',' in message.text:
                # делим ФИО и номер по запятой
                client_name = message.text.split(',')[0]
                client_number = message.text.split(',')[-1]
                # задаём имена переменных для параметров клиента из словаря
                file = user_data[message.chat.id]['file']
                doctor = user_data[message.chat.id]['doctor']
                page = user_data[message.chat.id]['page']
                t = user_data[message.chat.id]['time']
                # начинаем работу с данными в таблице, сверяем время и вносим запись
                column_a = page['A'][1:]
                column_b = page['B'][1:]
                for i in range(len(column_a)):
                    if str(column_a[i].value)[:-3] == t:
                        if column_b[i].value is None:
                            page['B' + str(column_b[i].row)] = client_name
                            page['C' + str(column_b[i].row)] = client_number
                            break
                        else:
                            bot.send_message(message.chat.id, 'Что-то пошло не так, вероятно, время уже занято. '
                                                              'Попробуйте ещё раз')
                            break
                file.save(doctor + '.xlsx')
                bot.send_message(message.chat.id, f'Ждём Вас {page.title} в {t} !')
                del user_data[message.chat.id]
                # очищаем словарик от временных данных по клиенту
            else:
                bot.send_message(message.chat.id,
                                 'Пожалуйста, введите своё имя (ФИО) и, через запятую - номер телефона')
        else:
            bot.send_message(message.chat.id, 'Пожалуйста, воспользуйтесь кнопками')


@bot.callback_query_handler(func=lambda call: True)
def callback_doctor(call):
    """Обработка диалога через инлайновые кнопки"""
    global user_data
    if call.message:
        # клиент выбрал доктора, передав данные в функцию
        files_names = [each.split('.xls')[0]
                       for each in os.listdir() if each.endswith('.xlsx')]
        if call.data in files_names:
            # работа с файлом, определение листов и их имён
            user_data[call.message.chat.id] = {}
            user_data[call.message.chat.id]['doctor'] = call.data
            doctor = user_data[call.message.chat.id]['doctor']
            file = openpyxl.load_workbook(str(doctor) + '.xlsx')
            user_data[call.message.chat.id]['file'] = file
            dates = file.sheetnames
            markup = types.InlineKeyboardMarkup()
            # цикл для создания кнопок с выбором даты (дата - это имя листа в табличном документе)
            for i in dates:
                markup.add(types.InlineKeyboardButton(str(i), callback_data=str(i)))
            bot.send_message(call.message.chat.id, f'На какую дату Вы хотели бы запланировать визит к {doctor}у?',
                             reply_markup=markup)
            bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                  text='Доктор:   ' + doctor, reply_markup=None)
        elif call.data in user_data[call.message.chat.id]['file'].sheetnames:
            # доктор (соответственно, файл) выбран, лист (соответственно, день) выбран, начинаем работать с таблицей
            page = user_data[call.message.chat.id]['file'][call.data]
            user_data[call.message.chat.id]['page'] = page
            markup = types.InlineKeyboardMarkup()
            column_a = page['A'][1:]
            column_b = page['B'][1:]
            time_list = []
            for i in range(len(column_a)):
                # цикл для создания кнопок с выбором времени, а также создания списка, который впоследствии пригодится
                if column_b[i].value is None:
                    t = str(column_a[i].value)[:-3]
                    time_list.append(t)
                else:
                    continue
                markup.add(types.InlineKeyboardButton(str(t), callback_data=str(t)))
            user_data[call.message.chat.id]['time_list'] = time_list
            bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                  text='Дата:   ' + call.data, reply_markup=None)
            bot.send_message(call.message.chat.id, 'Выберите доступное время:',
                             reply_markup=markup)
        elif call.data in user_data[call.message.chat.id]['time_list']:
            # дата выбрана, обрабатываем нажатие, клиент введет ФИО, номер, и запустится функция callback_doctor
            user_data[call.message.chat.id]['time'] = call.data
            bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                                  text='Время:   ' + call.data, reply_markup=None)
            bot.send_message(call.message.chat.id, text='Если всё верно, напишите, пожалуйста, ваше ФИО и '
                                                        'контактный телефон через запятую (пример: '
                                                        'Иванов Иван Иванович, +7**********). Если Вы хотите '
                                                        'скорректировать параметры, нажмите "Записаться к врачу" снова')


if __name__ == '__main__':
    bot.polling(none_stop=True)
